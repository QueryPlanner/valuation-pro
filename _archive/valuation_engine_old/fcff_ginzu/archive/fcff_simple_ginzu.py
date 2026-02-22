"""
ARCHIVED: Spreadsheet-mode adapter / CLI (parity checks)
=======================================================

This file is intentionally archived.

The core valuation engine is in `engine.py`.

This script remains useful for:
- Parity checks against the spreadsheet template CSV (path needs update)
- Optional diffs against cached Excel values (`valuation_engine/fcff_ginzu/spreadsheets/fcffsimpleginzu.xlsx`)
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import argparse
import csv
import json
import sys
import zipfile
import xml.etree.ElementTree as ET
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    # Prefer normal imports (works when run as `python -m archive.fcff_simple_ginzu`).
    from valuation_engine.fcff_ginzu import GinzuInputs, GinzuOutputs, InputError, compute_ginzu
except ModuleNotFoundError:
    # Allow running as a file: `python valuation_engine/fcff_ginzu/archive/fcff_simple_ginzu.py`
    project_root = Path(__file__).resolve().parents[3]
    sys.path.insert(0, str(project_root))
    from valuation_engine.fcff_ginzu import GinzuInputs, GinzuOutputs, InputError, compute_ginzu


def _as_bool_yes_no(value: str) -> bool:
    normalized = (value or "").strip().lower()
    if normalized in {"yes", "y", "true"}:
        return True
    if normalized in {"no", "n", "false"}:
        return False
    raise InputError(f"Expected Yes/No value, got: {value!r}")


def _as_float(value: Any) -> float:
    if value is None:
        raise InputError("Expected a number, got None")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "":
        raise InputError("Expected a number, got empty string")
    # Support simple “=a+b” style constants that appear in the CSV example.
    if text.startswith("=") and "+" in text and all(ch in "0123456789.+-=/ " for ch in text):
        text = text[1:].replace(" ", "")
        parts = text.split("+")
        return float(sum(float(p) for p in parts))
    return float(text)


def load_example_inputs_from_input_sheet_csv(
    input_sheet_csv_path: Path,
    *,
    wacc_initial: float,
    mature_market_erp: float = 0.0433,
) -> GinzuInputs:
    """
    Loads the *example* inputs from Input sheet.csv.

    Important:
    - The CSV stores formulas, not evaluated values.
    - For example-driven fields that are formulas in the CSV (like “Target margin = B27”),
      we follow the spreadsheet *intent* and derive them from the base numbers.
    """
    rows = _read_two_column_key_value_csv(input_sheet_csv_path)

    revenues_base = _as_float(rows["Revenues"])
    ebit_reported_base = _as_float(rows["Operating income or EBIT"])
    book_equity = _as_float(rows["Book value of equity"])
    book_debt = _as_float(rows["Book value of debt"])
    cash = _as_float(rows["Cash and Marketable Securities"])
    non_operating_assets = _as_float(rows["Cross holdings and other non-operating assets"])
    minority_interests = _as_float(rows["Minority interests"])
    shares_outstanding = _as_float(rows["Number of shares outstanding ="])
    stock_price = _as_float(rows["Current stock price ="])

    tax_rate_effective = _as_float(rows["Effective tax rate ="])
    tax_rate_marginal = _as_float(rows["Marginal tax rate ="])

    rev_growth_y1 = _as_float(rows["Revenue growth rate for next year"])

    riskfree_rate_now = _as_float(rows["Riskfree rate"])

    capitalize_rnd = _as_bool_yes_no(rows["Do you have R&D expenses to capitalize?"])
    capitalize_operating_leases = _as_bool_yes_no(rows["Do you have operating lease commitments?"])
    has_employee_options = _as_bool_yes_no(rows["Do you have employee options outstanding?"])

    rnd_asset = 0.0
    rnd_ebit_adjustment = 0.0
    if capitalize_rnd:
        rnd_asset, rnd_ebit_adjustment = load_rnd_adjustments_from_converter_csv(
            input_sheet_csv_path.parent / "R& D converter.csv"
        )

    def _get_optional_numeric_input(label: str) -> Optional[float]:
        """
        Return a float for a user-entered numeric cell, or None if:
        - the label is missing
        - the cell is blank
        - the cell is an Excel formula (starts with '=')

        This keeps `--example-from-csv` spreadsheet-faithful:
        - if the input sheet contains explicit values, we respect them
        - if the input sheet contains template formulas, we follow the intent instead
        """
        raw = rows.get(label)
        if raw is None:
            return None
        text = str(raw).strip()
        if text == "":
            return None
        if text.startswith("="):
            return None
        return _as_float(text)

    def _coerce_rate_0_to_1(label: str, value: float) -> float:
        """
        Guard against a common data-entry mistake when moving between spreadsheets/CSVs:
        using percent points (e.g. 14) instead of decimals (0.14).
        """
        if 0.0 <= value <= 1.0:
            return value
        if 1.0 < value <= 100.0:
            raise InputError(f"{label!r} looks like a percent (e.g. 14 for 14%). Use decimals (0.14). Got: {value}")
        return value

    # Many "value driver" cells are editable in practice, even if the template contains formulas.
    # We prefer explicit numeric inputs when present, and only fall back to template intent
    # when the CSV contains formulas (which are unevaluated in a CSV export).
    current_ebit_adjusted = ebit_reported_base + rnd_ebit_adjustment
    current_margin = current_ebit_adjusted / revenues_base

    rev_cagr_y2_5 = _get_optional_numeric_input("Compounded annual revenue growth rate - years 2-5 =")
    if rev_cagr_y2_5 is None:
        rev_cagr_y2_5 = rev_growth_y1
    rev_cagr_y2_5 = _coerce_rate_0_to_1("Compounded annual revenue growth rate - years 2-5 =", float(rev_cagr_y2_5))

    margin_y1 = _get_optional_numeric_input("Operating Margin for next year")
    if margin_y1 is None:
        margin_y1 = current_margin
    margin_y1 = _coerce_rate_0_to_1("Operating Margin for next year", float(margin_y1))

    margin_target = _get_optional_numeric_input("Target pre-tax operating margin =")
    if margin_target is None:
        margin_target = margin_y1
    margin_target = _coerce_rate_0_to_1("Target pre-tax operating margin =", float(margin_target))

    margin_convergence_year = int(_as_float(rows["Year of convergence for margin"]))

    sales_to_capital_1_5 = _get_optional_numeric_input("Sales to capital ratio  (for years 1-5)")
    sales_to_capital_6_10 = _get_optional_numeric_input("Sales to capital ratio (for years 6-10)")

    # Sales-to-capital ratios in the template are *formulas* by default, which pull the GLOBAL
    # industry average Sales/Capital via VLOOKUP on "Industry (Global)".
    #
    # If the CSV contains explicit numeric values, use them.
    if sales_to_capital_1_5 is None or sales_to_capital_6_10 is None:
        industry_global = (rows.get("Industry (Global)") or "").strip()
        if industry_global == "":
            raise InputError(
                "Missing 'Industry (Global)' in Input sheet.csv; cannot infer default sales-to-capital ratio"
            )
        fallback_sales_to_capital = _lookup_global_industry_sales_to_capital(
            industry_global=industry_global,
            industry_average_global_csv_path=input_sheet_csv_path.parent / "Industry Average Beta (Global).csv",
        )
        if sales_to_capital_1_5 is None:
            sales_to_capital_1_5 = fallback_sales_to_capital
        if sales_to_capital_6_10 is None:
            sales_to_capital_6_10 = fallback_sales_to_capital

    return GinzuInputs(
        revenues_base=revenues_base,
        ebit_reported_base=ebit_reported_base,
        book_equity=book_equity,
        book_debt=book_debt,
        cash=cash,
        non_operating_assets=non_operating_assets,
        minority_interests=minority_interests,
        shares_outstanding=shares_outstanding,
        stock_price=stock_price,
        rev_growth_y1=rev_growth_y1,
        rev_cagr_y2_5=rev_cagr_y2_5,
        margin_y1=margin_y1,
        margin_target=margin_target,
        margin_convergence_year=margin_convergence_year,
        sales_to_capital_1_5=float(sales_to_capital_1_5),
        sales_to_capital_6_10=float(sales_to_capital_6_10),
        riskfree_rate_now=riskfree_rate_now,
        wacc_initial=wacc_initial,
        tax_rate_effective=tax_rate_effective,
        tax_rate_marginal=tax_rate_marginal,
        capitalize_rnd=capitalize_rnd,
        capitalize_operating_leases=capitalize_operating_leases,
        has_employee_options=has_employee_options,
        rnd_asset=rnd_asset,
        rnd_ebit_adjustment=rnd_ebit_adjustment,
        mature_market_erp=mature_market_erp,
    )


def _lookup_global_industry_sales_to_capital(*, industry_global: str, industry_average_global_csv_path: Path) -> float:
    """
    Mirrors the template's VLOOKUP into `Industry Average Beta (Global)` for Sales/Capital.

    In the CSV, the column header is `Sales/Capital`.
    """
    if not industry_average_global_csv_path.exists():
        raise InputError(f"Missing required file: {industry_average_global_csv_path}")

    with industry_average_global_csv_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise InputError(f"Could not read headers from: {industry_average_global_csv_path}")

        expected_field = "Sales/Capital"
        if expected_field not in reader.fieldnames:
            raise InputError(
                f"Expected column {expected_field!r} in {industry_average_global_csv_path.name}, "
                f"found: {reader.fieldnames}"
            )

        for row in reader:
            row_industry = (row.get("Industry Name") or "").strip()
            if row_industry != industry_global:
                continue

            value = row.get(expected_field)
            if value is None:
                raise InputError(f"{expected_field!r} missing for industry {industry_global!r}")
            return _as_float(value)

    raise InputError(
        f"Industry {industry_global!r} not found in {industry_average_global_csv_path.name} "
        "(check spelling vs the CSV's 'Industry Name' column)"
    )


def load_rnd_adjustments_from_converter_csv(rnd_converter_csv_path: Path) -> Tuple[float, float]:
    """
    Compute the R&D capitalization adjustments from `R& D converter.csv`.

    We only need two outputs for the FCFF model:
    - rnd_asset: "Value of Research Asset" (used for completeness / potential extensions)
    - rnd_ebit_adjustment: "Adjustment to Operating Income" (add to reported EBIT)
    """
    if not rnd_converter_csv_path.exists():
        raise InputError(f"Missing required file for R&D capitalization: {rnd_converter_csv_path}")

    def _is_number_like(x: str) -> bool:
        try:
            _as_float(x)
            return True
        except Exception:
            return False

    amortization_years: Optional[int] = None
    current_year_rnd: Optional[float] = None
    past_year_expenses: List[float] = []

    with rnd_converter_csv_path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        raw_rows = list(reader)

    for row in raw_rows:
        if not row:
            continue
        label = (row[0] or "").strip()
        if label == "Over how many years do you want to amortize R&D expenses":
            candidate = row[5] if len(row) > 5 else ""
            amortization_years = int(_as_float(candidate))
            continue
        if label == "Enter the current year's R&D expense =":
            candidate = row[5] if len(row) > 5 else ""
            current_year_rnd = _as_float(candidate)
            continue

    if amortization_years is None or amortization_years <= 0:
        raise InputError("Could not read a valid R&D amortization period from R&D converter CSV")
    if amortization_years > 10:
        raise InputError(f"R&D amortization period must be <= 10 (spreadsheet constraint). Got: {amortization_years}")
    if current_year_rnd is None or current_year_rnd < 0:
        raise InputError("Could not read a valid current-year R&D expense from R&D converter CSV")

    table_start_idx: Optional[int] = None
    for i, row in enumerate(raw_rows):
        if len(row) >= 2 and (row[0] or "").strip() == "Year" and (row[1] or "").strip() == "R& D Expenses":
            table_start_idx = i + 1
            break

    if table_start_idx is None:
        raise InputError("Could not locate the 'Year, R& D Expenses' table in R&D converter CSV")

    needed_past_years = amortization_years - 1
    for row in raw_rows[table_start_idx:]:
        if len(past_year_expenses) >= needed_past_years:
            break
        if len(row) < 2:
            continue
        expense_cell = (row[1] or "").strip()
        if expense_cell == "":
            continue
        if not _is_number_like(expense_cell):
            continue
        past_year_expenses.append(_as_float(expense_cell))

    if len(past_year_expenses) != needed_past_years:
        raise InputError(
            "R&D capitalization is enabled, but the R&D converter sheet does not contain enough past-year expenses. "
            f"Need {needed_past_years} year(s) for an amortization period of {amortization_years}."
        )

    n = float(amortization_years)
    research_asset = float(current_year_rnd)
    amortization_this_year = 0.0
    for k, expense in enumerate(past_year_expenses, start=1):
        unamortized_fraction = (n - float(k)) / n
        research_asset += expense * unamortized_fraction
        amortization_this_year += expense / n

    ebit_adjustment = float(current_year_rnd) - amortization_this_year
    return research_asset, ebit_adjustment


def _read_two_column_key_value_csv(path: Path) -> Dict[str, str]:
    """
    Reads Input sheet.csv as label->value using the first two columns:
    - column 0: label
    - column 1: value

    Skips empty labels.
    """
    rows: Dict[str, str] = {}
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row:
                continue
            label = (row[0] or "").strip()
            if label == "":
                continue
            value = row[1] if len(row) > 1 else ""
            rows[label] = value
    return rows


def _format_money(x: float) -> str:
    return f"{x:,.2f}"


def _print_summary(outputs: GinzuOutputs) -> None:
    print("FCFF Simple Ginzu — Key Outputs")
    print(f"- PV (CF over next 10 years): {_format_money(outputs.pv_10y)}")
    print(f"- PV(Terminal value): {_format_money(outputs.pv_terminal_value)}")
    print(f"- Sum of PV: {_format_money(outputs.pv_sum)}")
    print(f"- Value of operating assets: {_format_money(outputs.value_of_operating_assets)}")
    print(f"- Value of equity in common stock: {_format_money(outputs.value_of_equity_common)}")
    print(f"- Estimated value /share: {_format_money(outputs.estimated_value_per_share)}")
    print(f"- Price as % of value: {outputs.price_as_percent_of_value:.4f}")


def _load_inputs_from_json(path: Path) -> GinzuInputs:
    data = json.loads(path.read_text(encoding="utf-8"))
    return GinzuInputs(**data)


def _diff_against_xlsx(
    *,
    xlsx_path: Path,
    expected_sheet_name: str,
    cell_map: Dict[str, str],
    outputs: GinzuOutputs,
    tolerance: float,
) -> None:
    """
    Optional: compare computed outputs to cached Excel values.
    """
    try:  # pragma: no cover
        from openpyxl import load_workbook

        wb = load_workbook(filename=str(xlsx_path), data_only=True)
        if expected_sheet_name not in wb.sheetnames:
            raise InputError(f"Sheet {expected_sheet_name!r} not found in workbook. Found: {wb.sheetnames}")
        ws = wb[expected_sheet_name]

        def read_expected_cell(address: str) -> Optional[float]:
            value = ws[address].value
            if value is None:
                return None
            return float(value)

    except Exception:
        expected_cells = _read_numeric_cells_from_xlsx(
            xlsx_path=xlsx_path,
            sheet_name=expected_sheet_name,
            cell_addresses=set(cell_map.values()),
        )

        def read_expected_cell(address: str) -> Optional[float]:
            return expected_cells.get(address)

    computed: Dict[str, float] = {
        "PV_10y": outputs.pv_10y,
        "PV_TerminalValue": outputs.pv_terminal_value,
        "Value_OperatingAssets": outputs.value_of_operating_assets,
        "Value_EquityCommon": outputs.value_of_equity_common,
        "Value_PerShare": outputs.estimated_value_per_share,
    }

    print("\nComparison vs Excel cached values (data_only=True)")
    print(f"(tolerance = {tolerance})")
    for label, address in cell_map.items():
        expected_value = read_expected_cell(address)
        computed_value = computed.get(label)
        if computed_value is None:
            continue
        if expected_value is None:
            print(f"- {label}: Excel {address} is empty; computed={computed_value}")
            continue
        diff = computed_value - expected_value
        ok = abs(diff) <= tolerance
        status = "OK" if ok else "DIFF"
        print(f"- {label}: {status}  computed={computed_value:.6f}  excel={expected_value:.6f}  diff={diff:.6f}")


def _read_numeric_cells_from_xlsx(*, xlsx_path: Path, sheet_name: str, cell_addresses: set[str]) -> Dict[str, float]:
    """
    Reads cached *numeric* values from an .xlsx for a specific sheet + a small set of cell addresses.
    """
    if not xlsx_path.exists():
        raise InputError(f"xlsx file not found: {xlsx_path}")
    if not cell_addresses:
        return {}

    sheet_xml_path = _find_sheet_xml_path_in_xlsx(xlsx_path=xlsx_path, sheet_name=sheet_name)
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        raw_sheet_xml = zf.read(sheet_xml_path)

    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(raw_sheet_xml)

    wanted = set(cell_addresses)
    found: Dict[str, float] = {}

    for cell in root.findall(".//main:c", ns):
        address = cell.get("r")
        if not address or address not in wanted:
            continue

        value_node = cell.find("main:v", ns)
        if value_node is None or value_node.text is None:
            continue

        try:
            found[address] = float(value_node.text)
        except ValueError:
            continue

        if len(found) == len(wanted):
            break

    return found


def _find_sheet_xml_path_in_xlsx(*, xlsx_path: Path, sheet_name: str) -> str:
    """
    Resolves a sheet name to its worksheet XML path inside the .xlsx zip.
    """
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        workbook_xml = zf.read("xl/workbook.xml")
        rels_xml = zf.read("xl/_rels/workbook.xml.rels")

    wb_ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    rel_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

    wb_root = ET.fromstring(workbook_xml)
    rels_root = ET.fromstring(rels_xml)

    target_rel_id: Optional[str] = None
    for sheet in wb_root.findall(".//main:sheets/main:sheet", wb_ns):
        if sheet.get("name") == sheet_name:
            target_rel_id = sheet.get(f"{{{wb_ns['r']}}}id")
            break

    if not target_rel_id:
        raise InputError(f"Sheet {sheet_name!r} not found in workbook.xml")

    target_path: Optional[str] = None
    for rel in rels_root.findall(".//rel:Relationship", rel_ns):
        if rel.get("Id") == target_rel_id:
            target_path = rel.get("Target")
            break

    if not target_path:
        raise InputError(f"Could not resolve relationship {target_rel_id!r} for sheet {sheet_name!r}")

    return f"xl/{target_path.lstrip('/')}"


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="FCFF Simple Ginzu valuation calculator (archived spreadsheet-mode CLI)")
    parser.add_argument(
        "--inputs-json",
        type=Path,
        help="Path to a JSON file containing GinzuInputs (as a dict).",
    )
    parser.add_argument(
        "--example-from-csv",
        action="store_true",
        help="Load the example company inputs from fcffsimpleginzu-formulas/Input sheet.csv (WACC still required).",
    )
    parser.add_argument(
        "--wacc-initial",
        type=float,
        required=False,
        help="Initial WACC (required for --example-from-csv).",
    )
    parser.add_argument(
        "--compare-xlsx",
        type=Path,
        help="Optional: compare key outputs to cached values in an .xlsx file using openpyxl (requires openpyxl).",
    )
    parser.add_argument(
        "--compare-xlsx-sheet",
        type=str,
        default="Valuation output",
        help="Sheet name in the .xlsx to read for comparison (default: 'Valuation output').",
    )
    parser.add_argument(
        "--tolerance",
        type=float,
        default=1e-4,
        help="Numeric tolerance for comparisons (default: 1e-4).",
    )

    args = parser.parse_args(argv)

    inputs: Optional[GinzuInputs] = None
    if args.inputs_json:
        inputs = _load_inputs_from_json(args.inputs_json)

    if args.example_from_csv:
        if args.wacc_initial is None:
            raise InputError("--wacc-initial is required when using --example-from-csv")
        inputs = load_example_inputs_from_input_sheet_csv(
            Path(__file__).resolve().parent.parent / "fcffsimpleginzu-formulas" / "Input sheet.csv",
            wacc_initial=args.wacc_initial,
        )

    if inputs is None:
        raise InputError("Provide either --inputs-json or --example-from-csv.")

    outputs = compute_ginzu(inputs)
    _print_summary(outputs)

    if args.compare_xlsx:
        cell_map = {
            "PV_10y": "B20",
            "PV_TerminalValue": "B19",
            "Value_OperatingAssets": "B24",
            "Value_EquityCommon": "B31",
            "Value_PerShare": "B33",
        }
        _diff_against_xlsx(
            xlsx_path=args.compare_xlsx,
            expected_sheet_name=args.compare_xlsx_sheet,
            cell_map=cell_map,
            outputs=outputs,
            tolerance=args.tolerance,
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())


