
import openpyxl
import sys
import os

def check_excel_truth(file_path, sheet_name="Valuation output"):
    print(f"Loading {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return

    ws = wb[sheet_name]
    
    # Mapping based on Valuation output.csv structure
    # Row 33 Column B is Estimated Value / Share
    # In openpyxl, rows are 1-based, columns are 1-based (A=1, B=2)
    
    # Key cells to check in "Valuation output"
    # B33: Estimated value / share
    # B31: Value of equity in common stock
    # B21: Value of operating assets
    # M12: Terminal cost of capital
    # B3: Revenues (Base)
    # B27: Cash
    # B25: Debt
    
    mapping = {
        "Value per Share": "B33",
        "Value of Equity": "B31",
        "Value of Op Assets": "B21",
        "Terminal WACC": "M12",
        "Base Revenue": "B3",
        "Base Margin (B4)": "B4",
        "Base EBIT (B5)": "B5",
        "Adjusted Debt": "B25",
        "Adjusted Cash": "B27",
    }

    print(f"\n--- Truth from {sheet_name} ---")
    for name, cell_ref in mapping.items():
        val = ws[cell_ref].value
        print(f"{name} ({cell_ref}): {val}")

    # Check Input sheet for confirmation
    input_sheet = wb["Input sheet"]
    print(f"\n--- Key Inputs from 'Input sheet' ---")
    input_mapping = {
        "Revenues (B11)": "B11",
        "Reported EBIT (B12)": "B12",
        "Capitalize R&D? (B16)": "B16",
        "Capitalize Leases? (B17)": "B17",
        "Growth Year 1 (B26)": "B26",
        "Margin Year 1 (I26)": "I26", 
        "Target Margin (B29)": "B29",
    }
    
    for name, cell_ref in input_mapping.items():
        val = input_sheet[cell_ref].value
        print(f"{name}: {val}")

    # Check R&D Sheet if available
    if "R& D converter" in wb.sheetnames:
        rd_sheet = wb["R& D converter"]
        print(f"\n--- Key R&D Outputs ---")
        # D39: Adjustment to Operating Income
        # D37: Amortization
        # F7: Current Year R&D
        # F6: Amortization Years
        rd_mapping = {
            "Amortization Years (F6)": "F6",
            "Current R&D (F7)": "F7",
            "Amortization (D37)": "D37",
            "Adjustment to EBIT (D39)": "D39",
            "Value of Research Asset (D35)": "D35",
        }
        for name, cell_ref in rd_mapping.items():
            val = rd_sheet[cell_ref].value
            print(f"{name}: {val}")

if __name__ == "__main__":
    # Default to AMZN, but can be easily switched to KO
    company = "amzn" # or "ko"
    filename = "amzn_valuation.xlsx" if company == "amzn" else "ko_valuation.xlsx"
    base_path = os.path.join(os.path.dirname(__file__), "..", "reference_models", company, filename)
    check_excel_truth(base_path)
