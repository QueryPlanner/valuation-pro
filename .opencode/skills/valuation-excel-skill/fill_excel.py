import json
import argparse
import openpyxl
from datetime import datetime
import yfinance as yf

def fill_valuation_excel(company_name, inputs_json_path, template_path, output_path, price=None, rf_rate=None, erp=0.055, ticker=None):
    # If ticker is provided, try to fetch dynamic market data
    if ticker:
        try:
            stock = yf.Ticker(ticker)
            if price is None:
                price = stock.fast_info.last_price
            if rf_rate is None:
                tnx = yf.Ticker("^TNX")
                rf_rate = tnx.fast_info.last_price / 100
        except Exception as e:
            print(f"Failed to fetch market data via yfinance: {e}")
    
    # Fallbacks if yfinance failed or no args provided
    price = price if price is not None else 100.0
    rf_rate = rf_rate if rf_rate is not None else 0.045
    # Load valuation inputs
    with open(inputs_json_path, "r") as f:
        inputs = json.load(f)

    if isinstance(inputs, list):
        inputs = inputs[0]
    
    fin_data = inputs.get("financial_data", {})
    single_metrics = inputs.get("single_value_metrics", {})
    cost_of_capital = inputs.get("cost_of_capital_inputs", {})

    r_and_d_details = inputs.get("r_and_d_details", {})
    industry_name = r_and_d_details.get("industry_name", "Drugs (Pharmaceutical)")

    wb = openpyxl.load_workbook(template_path)
    ws = wb['Input sheet']

    # Helper to write to cell safely
    def write_cell(cell, value):
        ws[cell] = value

    # Basic info
    write_cell('B3', datetime.today().strftime('%m/%d/%Y'))
    write_cell('B4', company_name)
    write_cell('B7', "United States") # Defaulting to US, can be dynamic
    write_cell('B8', industry_name)
    write_cell('B9', industry_name)

    # Fill LTM numbers
    revenues_ltm = fin_data.get("Revenues", {}).get("Most_Recent_12_months", 0.0)
    write_cell('B11', revenues_ltm) 
    last_10k_rev = fin_data.get("Revenues", {}).get("Last_10K_before_LTM", 0.0)
    write_cell('C11', last_10k_rev) 
    write_cell('D11', single_metrics.get("Years_since_last_10K", 1.0))      

    ebit_ltm = fin_data.get("Operating_income_or_EBIT", {}).get("Most_Recent_12_months", 0.0)
    write_cell('B12', ebit_ltm)  
    write_cell('C12', fin_data.get("Operating_income_or_EBIT", {}).get("Last_10K_before_LTM", 0.0))  

    write_cell('B13', fin_data.get("Interest_expense", {}).get("Most_Recent_12_months", 0.0))   
    write_cell('C13', fin_data.get("Interest_expense", {}).get("Last_10K_before_LTM", 0.0))   

    write_cell('B14', fin_data.get("Book_value_of_equity", {}).get("Most_Recent_12_months", 0.0)) 
    write_cell('C14', fin_data.get("Book_value_of_equity", {}).get("Last_10K_before_LTM", 0.0)) 

    debt_ltm = fin_data.get("Book_value_of_debt", {}).get("Most_Recent_12_months", 0.0)
    write_cell('B15', debt_ltm)  
    write_cell('C15', fin_data.get("Book_value_of_debt", {}).get("Last_10K_before_LTM", 0.0))  

    write_cell('B16', 'No')     
    write_cell('B17', 'No')     

    cash_ltm = fin_data.get("Cash_and_Marketable_Securities", {}).get("Most_Recent_12_months", 0.0)
    write_cell('B18', cash_ltm)  
    write_cell('C18', fin_data.get("Cash_and_Marketable_Securities", {}).get("Last_10K_before_LTM", 0.0))  

    write_cell('B19', fin_data.get("Cross_holdings_and_other_non_operating_assets", {}).get("Most_Recent_12_months", 0.0))   
    write_cell('C19', fin_data.get("Cross_holdings_and_other_non_operating_assets", {}).get("Last_10K_before_LTM", 0.0))   

    write_cell('B20', fin_data.get("Minority_interests", {}).get("Most_Recent_12_months", 0.0))   
    write_cell('C20', fin_data.get("Minority_interests", {}).get("Last_10K_before_LTM", 0.0))      

    shares_out = single_metrics.get("Number_of_shares_outstanding", 0.0)
    write_cell('B21', shares_out)   
    
    # Fill stock price dynamically
    write_cell('B22', price)  

    write_cell('B23', single_metrics.get("Effective_tax_rate", 0.25))   
    write_cell('B24', 0.25)     

    base_growth = (revenues_ltm / last_10k_rev) - 1.0 if last_10k_rev > 0 else 0.05
    current_margin = ebit_ltm / revenues_ltm if revenues_ltm > 0 else 0.15

    # Value drivers
    write_cell('B26', base_growth)   
    write_cell('B27', current_margin)   
    write_cell('B28', base_growth)   
    write_cell('B29', current_margin)   
    write_cell('B30', 5.0)      
    write_cell('B31', 1.0)      
    write_cell('B32', 1.0)      

    # Riskfree rate & Cost of Capital
    write_cell('B34', rf_rate)     
    write_cell('B35', "='Cost of capital worksheet'!B13")

    # Employee options
    write_cell('B37', 'No')

    # Stories to Numbers Worksheet: Provide user guidance instead of automating a story
    if 'Stories to Numbers' in wb.sheetnames:
        ws_story = wb['Stories to Numbers']
        ws_story['A2'] = "[Insert Your Company Story Title Here]"
        ws_story['A3'] = "GUIDANCE: Read the 'Management Discussion & Analysis' (MD&A) section in the Annual Report. What is the company's core business model? Who are their competitors? Are they a disruptor or an incumbent? Write a narrative here tying their business strategy to their ability to generate cash flows, drive growth, and manage risk."
        ws_story['G9'] = "GUIDANCE: Link to Growth Story. Look at historical revenue trends and management's future guidance. Why will they grow at this rate?"
        ws_story['G10'] = "GUIDANCE: Link to Profitability. Are they cutting costs or scaling? Compare current margins to industry averages to set a realistic target."
        ws_story['G11'] = "GUIDANCE: Check the effective tax rate in the income statement vs the marginal corporate tax rate of their home country."
        ws_story['G12'] = "GUIDANCE: Link to Capital Efficiency. How much reinvestment (CapEx, R&D, Acquisitions) is needed to drive the growth story above?"
        ws_story['G13'] = "GUIDANCE: Link to Competitive Advantage (Moat). A high ROC requires strong barriers to entry. Do they have them?"
        ws_story['G14'] = "GUIDANCE: Link to Risk Profile. Look at their debt load, geographical exposure, and business cyclicality to justify the Cost of Capital."

    # Cost of Capital Worksheet
    if 'Cost of capital worksheet' in wb.sheetnames:
        ws_coc = wb['Cost of capital worksheet']

        def write_coc_cell(cell, value):
            ws_coc[cell] = value

        write_coc_cell('B11', 'Detailed')
        write_coc_cell('B18', shares_out)
        write_coc_cell('B19', price)
        
        by_business = inputs.get("revenue_splits", {}).get("by_business", {})
        is_multibusiness = len(by_business) > 1
        
        if is_multibusiness:
            write_coc_cell('B21', 'Multibusiness(US)')
            
            # Clear existing dummy entries in G36:H47
            for i in range(36, 48):
                write_coc_cell(f'G{i}', None)
                write_coc_cell(f'H{i}', None)
                
            # Populate actual businesses
            row = 36
            for biz_name, rev in by_business.items():
                if row <= 47:
                    write_coc_cell(f'G{row}', biz_name)
                    write_coc_cell(f'H{row}', rev)
                    row += 1
        else:
            write_coc_cell('B21', 'Single Business(Global)')

        write_coc_cell('B24', rf_rate)
        write_coc_cell('B25', 'Direct input')
        write_coc_cell('B26', erp) 
        write_coc_cell('B30', debt_ltm)
        write_coc_cell('B31', fin_data.get("Interest_expense", {}).get("Most_Recent_12_months", 0.0))
        write_coc_cell('B32', cost_of_capital.get("average_maturity_of_debt_years", 5.0))
        write_coc_cell('B33', 'Synthetic rating')
        write_coc_cell('B36', 1) 
        write_coc_cell('B38', single_metrics.get("Effective_tax_rate", 0.25))

    wb.save(output_path)
    print(f"Successfully populated {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Fill Valuation Excel")
    parser.add_argument("--company", type=str, required=True)
    parser.add_argument("--inputs", type=str, required=True)
    parser.add_argument("--template", type=str, default="template.xlsx")
    parser.add_argument("--output", type=str, required=True)
    parser.add_argument("--price", type=float, default=None, help="Current stock price")
    parser.add_argument("--rf_rate", type=float, default=None, help="Risk free rate (e.g. 0.044 for 4.4%)")
    parser.add_argument("--erp", type=float, default=0.055, help="Equity Risk Premium (e.g. 0.055 for 5.5%)")
    parser.add_argument("--ticker", type=str, default=None, help="Yahoo Finance ticker symbol to automatically fetch price and risk-free rate")
    args = parser.parse_args()
    
    fill_valuation_excel(args.company, args.inputs, args.template, args.output, args.price, args.rf_rate, args.erp, args.ticker)